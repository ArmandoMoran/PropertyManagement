import openpyxl
import pyodbc
import datetime

# --- Configuration ---
EXCEL_PATH = r'c:\SWAM DOCUMENTS\PROPERTIES\CURRENT Properties\1 MASTER\MASTER Property INFO Spreadsheet.xlsx'
CONN_STR = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER=localhost;DATABASE=PropertyManagement;Trusted_Connection=yes;'

# --- Open workbook ---
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
conn = pyodbc.connect(CONN_STR)
cursor = conn.cursor()

# ============================================================
# 1. INSERT PROPERTIES
# ============================================================
print("Inserting Properties...")
ws = wb['Properties']
property_map = {}  # address_key -> PropertyId

for row in ws.iter_rows(min_row=2, max_row=24, values_only=True):
    full_address = row[0]
    if not full_address:
        continue
    owner = row[1]
    prop_type = row[2]
    units_val = row[3]
    try:
        units = int(units_val) if units_val else None
    except:
        units = None

    cursor.execute(
        "INSERT INTO Properties (FullAddress, Owner, PropertyType, Units) OUTPUT INSERTED.PropertyId VALUES (?, ?, ?, ?)",
        full_address.strip(), 
        owner.strip() if owner else None,
        prop_type.strip() if prop_type else None,
        units
    )
    pid = int(cursor.fetchone()[0])
    
    # Build lookup key from the first word(s) of the address for matching across sheets
    addr_key = full_address.strip().split(',')[0].strip()
    property_map[addr_key] = pid
    # Also store shorter name keys for matching
    words = addr_key.split()
    if len(words) >= 2:
        short_key = words[0] + ' ' + words[1]
        property_map[short_key] = pid
        # Just the street name (first word or two)
        property_map[words[0]] = pid

conn.commit()
print(f"  Inserted {len(set(property_map.values()))} properties")

# Helper to match property address to PropertyId
def find_property_id(address):
    if not address:
        return None
    addr = address.strip()
    
    # Try exact match first
    if addr in property_map:
        return property_map[addr]
    
    # Try match on first part before comma
    addr_key = addr.split(',')[0].strip()
    if addr_key in property_map:
        return property_map[addr_key]
    
    # Try first two words
    words = addr_key.split()
    if len(words) >= 2:
        key2 = words[0] + ' ' + words[1]
        if key2 in property_map:
            return property_map[key2]
    
    # Try first word only
    if words:
        if words[0] in property_map:
            return property_map[words[0]]
    
    # Fuzzy: check if any key starts with first word
    first_word = words[0].lower() if words else ''
    for k, v in property_map.items():
        if k.lower().startswith(first_word) and first_word:
            return v
    
    return None

# ============================================================
# 2. INSERT LENDERS
# ============================================================
print("Inserting Lenders...")
ws = wb['Properties']
lender_count = 0

for row in ws.iter_rows(min_row=2, max_row=24, values_only=True):
    full_address = row[0]
    if not full_address:
        continue
    pid = find_property_id(full_address)
    if not pid:
        print(f"  WARNING: No property match for lender: {full_address}")
        continue
    
    lender_name = row[4]  # Lender
    lender_url = row[5]   # Lender Url
    user_id = row[6]      # User Id
    mortgage_num = row[7]  # Mortgage #
    monthly_pmt = row[8]   # Monthly Payment

    # Skip if lender is n/a
    if lender_name and str(lender_name).strip().lower() != 'n/a':
        pmt = None
        if monthly_pmt and str(monthly_pmt).strip().lower() not in ('na', 'n/a', ''):
            try:
                pmt = float(monthly_pmt)
            except:
                pmt = None
        
        cursor.execute(
            "INSERT INTO Lenders (PropertyId, LenderName, LenderUrl, UserId, MortgageNumber, MonthlyPayment, EffectiveDate) VALUES (?, ?, ?, ?, ?, ?, ?)",
            pid,
            str(lender_name).strip() if lender_name else None,
            str(lender_url).strip() if lender_url and str(lender_url).strip().lower() != 'n/a' else None,
            str(user_id).strip() if user_id else None,
            str(mortgage_num).strip() if mortgage_num and str(mortgage_num).strip().lower() != 'n/a' else None,
            pmt,
            datetime.date(2025, 10, 1)  # current effective date
        )
        lender_count += 1

conn.commit()
print(f"  Inserted {lender_count} lenders")

# ============================================================
# 3. INSERT PRINCIPAL BALANCE HISTORY
# ============================================================
print("Inserting Principal Balance History...")
ws = wb['Properties']
balance_count = 0

# Column mapping: J=Feb2022, K=Aug2024, L=Oct2025
balance_dates = {
    9: datetime.date(2022, 2, 1),   # col J (index 9)
    10: datetime.date(2024, 8, 1),  # col K (index 10)
    11: datetime.date(2025, 10, 1), # col L (index 11)
}

for row in ws.iter_rows(min_row=2, max_row=24, values_only=True):
    full_address = row[0]
    if not full_address:
        continue
    pid = find_property_id(full_address)
    if not pid:
        continue
    
    for col_idx, snap_date in balance_dates.items():
        val = row[col_idx]
        if val and str(val).strip().lower() not in ('n/a', 'na', ''):
            try:
                balance = float(val)
                cursor.execute(
                    "INSERT INTO PrincipalBalanceHistory (PropertyId, SnapshotDate, PrincipalBalance) VALUES (?, ?, ?)",
                    pid, snap_date, balance
                )
                balance_count += 1
            except (ValueError, TypeError):
                pass

conn.commit()
print(f"  Inserted {balance_count} balance records")

# ============================================================
# 4. INSERT HOA
# ============================================================
print("Inserting HOA records...")
ws = wb['HOA Detail']
hoa_count = 0

for row in ws.iter_rows(min_row=2, max_row=24, values_only=True):
    full_address = row[0]
    hoa_name = row[1]
    if not full_address:
        continue
    if not hoa_name or str(hoa_name).strip().lower() == 'na':
        continue

    pid = find_property_id(full_address)
    if not pid:
        print(f"  WARNING: No property match for HOA: {full_address}")
        continue
    
    account = str(row[2]).strip() if row[2] else None
    mgmt_company = str(row[3]).strip() if row[3] else None
    url = str(row[4]).strip() if row[4] else None
    phone = str(row[5]).strip() if row[5] else None
    contact = str(row[6]).strip() if row[6] else None
    frequency = str(row[7]).strip() if row[7] else None
    amount = None
    if row[8]:
        try:
            amount = float(row[8])
        except:
            pass
    
    cursor.execute(
        "INSERT INTO HOA (PropertyId, HOAName, AccountNumber, ManagementCompany, Url, Phone, Contact, PaymentFrequency, PaymentAmount, EffectiveYear) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        pid,
        str(hoa_name).strip(),
        account,
        mgmt_company,
        url,
        phone,
        contact,
        frequency,
        amount,
        2025  # current year
    )
    hoa_count += 1

conn.commit()
print(f"  Inserted {hoa_count} HOA records")

# Also insert historical HOA amounts from rows 30+ (2025/2026 yearly detail)
print("Inserting HOA historical amounts...")
hoa_hist_count = 0
for row in ws.iter_rows(min_row=31, max_row=42, values_only=True):
    full_address = row[0]
    if not full_address or not isinstance(full_address, str):
        continue
    frequency = row[1]
    amount_2025 = row[2]
    amount_2026 = row[6] if len(row) > 6 else None
    
    pid = find_property_id(full_address)
    if not pid:
        continue
    
    # Check if we already have a 2025 record for this property
    # Insert 2026 records if available
    if amount_2026:
        try:
            amt = float(amount_2026)
            cursor.execute(
                "INSERT INTO HOA (PropertyId, HOAName, PaymentFrequency, PaymentAmount, EffectiveYear) SELECT PropertyId, HOAName, PaymentFrequency, ?, 2026 FROM HOA WHERE PropertyId = ? AND EffectiveYear = 2025",
                amt, pid
            )
            hoa_hist_count += 1
        except:
            pass

conn.commit()
print(f"  Inserted {hoa_hist_count} HOA historical records")

# ============================================================
# 5. INSERT INSURANCE
# ============================================================
print("Inserting Insurance records...")
ws = wb['Insurance Detail2']
insurance_count = 0
premium_count = 0

for row in ws.iter_rows(min_row=2, max_row=24, values_only=True):
    full_address = row[0]
    if not full_address or str(full_address).strip().upper() == 'TOTALS':
        continue
    
    pid = find_property_id(full_address)
    if not pid:
        print(f"  WARNING: No property match for Insurance: {full_address}")
        continue
    
    carrier = str(row[1]).strip() if row[1] else None
    policy_no = str(row[2]).strip() if row[2] else None
    
    renewal_date = None
    if row[3] and isinstance(row[3], datetime.datetime):
        renewal_date = row[3].date()
    
    who_pays = str(row[5]).strip() if row[5] else None
    
    cursor.execute(
        "INSERT INTO Insurance (PropertyId, Carrier, PolicyNumber, RenewalDate, WhoPays) OUTPUT INSERTED.InsuranceId VALUES (?, ?, ?, ?, ?)",
        pid, carrier, policy_no, renewal_date, who_pays
    )
    ins_id = int(cursor.fetchone()[0])
    insurance_count += 1
    
    # Insert premium history for years 2021-2025
    # Columns: G=2021, H=2022, I=YOY, J=2023, K=YOY, L=2024, M=YOY, N=2025, O=YOY
    year_cols = [
        (2021, 6, None),    # col G, no YOY
        (2022, 7, 8),       # col H, YOY in col I
        (2023, 9, 10),      # col J, YOY in col K
        (2024, 11, 12),     # col L, YOY in col M
        (2025, 13, 14),     # col N, YOY in col O
    ]
    
    for year, prem_idx, yoy_idx in year_cols:
        if prem_idx < len(row) and row[prem_idx]:
            prem_val = row[prem_idx]
            if str(prem_val).strip().lower() in ('na', 'n/a', ''):
                continue
            try:
                premium = float(prem_val)
            except (ValueError, TypeError):
                continue
            
            yoy = None
            if yoy_idx and yoy_idx < len(row) and row[yoy_idx]:
                yoy_val = row[yoy_idx]
                if str(yoy_val).strip().lower() not in ('na', 'n/a', ''):
                    try:
                        yoy = round(float(yoy_val) * 100, 2)  # Convert to percentage
                    except:
                        pass
            
            cursor.execute(
                "INSERT INTO InsurancePremiumHistory (InsuranceId, PolicyYear, AnnualPremium, YOYPercentChange) VALUES (?, ?, ?, ?)",
                ins_id, year, premium, yoy
            )
            premium_count += 1

conn.commit()
print(f"  Inserted {insurance_count} insurance policies, {premium_count} premium records")

# ============================================================
# 6. INSERT PROPERTY HISTORY
# ============================================================
print("Inserting Property History...")
ws = wb['Property History']
history_count = 0
unmatched = set()

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    event_date = row[0]
    prop_name = row[1]
    description = row[2]
    notes = row[3]
    
    if not prop_name and not description:
        continue
    
    # Try to match property name to PropertyId
    pid = None
    if prop_name:
        pname = str(prop_name).strip()
        pid = find_property_id(pname)
        if not pid:
            # Try with common short names
            name_parts = pname.split()
            if name_parts:
                for k, v in property_map.items():
                    if name_parts[0].lower() in k.lower():
                        pid = v
                        break
    
    edate = None
    if event_date and isinstance(event_date, datetime.datetime):
        edate = event_date.date()
    
    if not pid and prop_name:
        unmatched.add(str(prop_name).strip())
    
    cursor.execute(
        "INSERT INTO PropertyHistory (PropertyId, EventDate, PropertyName, Description, Notes) VALUES (?, ?, ?, ?, ?)",
        pid,
        edate,
        str(prop_name).strip() if prop_name else None,
        str(description).strip() if description else None,
        str(notes).strip() if notes else None
    )
    history_count += 1

conn.commit()
print(f"  Inserted {history_count} history records")
if unmatched:
    print(f"  Unmatched property names: {unmatched}")

# --- Cleanup ---
cursor.close()
conn.close()
wb.close()

print("\n=== IMPORT COMPLETE ===")
