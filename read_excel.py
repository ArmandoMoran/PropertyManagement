import openpyxl

wb = openpyxl.load_workbook(r'c:\SWAM DOCUMENTS\PROPERTIES\CURRENT Properties\1 MASTER\MASTER Property INFO Spreadsheet.xlsx', data_only=True)

sheets_to_inspect = ['Properties', 'HOA Detail', 'Insurance Detail2', 'Property History']

for sheet_name in sheets_to_inspect:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
    print(f"{'='*80}")
    
    # Print first 15 rows to understand structure
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=False), 1):
        values = [(cell.column_letter, cell.value) for cell in row if cell.value is not None]
        if values:
            print(f"Row {row_idx}: {values}")
