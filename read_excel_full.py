import openpyxl
import json

wb = openpyxl.load_workbook(r'c:\SWAM DOCUMENTS\PROPERTIES\CURRENT Properties\1 MASTER\MASTER Property INFO Spreadsheet.xlsx', data_only=True)

# Full Properties sheet
ws = wb['Properties']
print("=== PROPERTIES (ALL ROWS) ===")
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
    values = [(cell.column_letter, cell.value) for cell in row if cell.value is not None]
    if values:
        print(f"Row {row_idx}: {values}")

# Full HOA Detail sheet
ws = wb['HOA Detail']
print("\n=== HOA DETAIL (ALL ROWS) ===")
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
    values = [(cell.column_letter, cell.value) for cell in row if cell.value is not None]
    if values:
        print(f"Row {row_idx}: {values}")

# Full Insurance Detail2 sheet
ws = wb['Insurance Detail2']
print("\n=== INSURANCE DETAIL (ALL ROWS) ===")
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
    values = [(cell.column_letter, cell.value) for cell in row if cell.value is not None]
    if values:
        print(f"Row {row_idx}: {values}")

# Property History count
ws = wb['Property History']
print(f"\n=== PROPERTY HISTORY: {ws.max_row} rows ===")
